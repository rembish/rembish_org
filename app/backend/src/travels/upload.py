from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..auth.session import get_admin_user
from ..database import get_db
from ..models import NMRegion, User
from .models import UploadResult

router = APIRouter()


def extract_country(region_name: str) -> str:
    """Extract country/territory prefix from region name."""
    if region_name.startswith("Timor-Leste"):
        return "Timor-Leste"
    if " – " in region_name:
        return region_name.split(" – ")[0].strip()
    if " - " in region_name:
        return region_name.split(" - ")[0].strip()
    return region_name


@router.post("/upload-nm", response_model=UploadResult)
async def upload_nm_regions(
    file: Annotated[UploadFile, File()],
    admin: Annotated[User, Depends(get_admin_user)],
    db: Session = Depends(get_db),
) -> UploadResult:
    """Upload NomadMania regions XLSX file (admin only)."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be an .xlsx file")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid XLSX file")

    ws = wb.active

    # Parse and validate all data BEFORE touching the database
    parsed_regions: list[dict] = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        name = row[0]
        if not name:
            continue

        try:
            parsed_regions.append(
                {
                    "name": str(name),
                    "country": extract_country(str(name)),
                    "visited": row[1] == 1,
                    "first_visited_year": int(row[2]) if row[2] else None,
                    "last_visited_year": int(row[3]) if row[3] else None,
                }
            )
        except (ValueError, TypeError) as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid data in row {row_num}: {e}",
            )

    if not parsed_regions:
        raise HTTPException(status_code=400, detail="No valid regions found in file")

    # All data validated - now safe to update database
    try:
        db.query(NMRegion).delete()

        visited_count = 0
        for region_data in parsed_regions:
            region = NMRegion(**region_data)
            db.add(region)
            if region_data["visited"]:
                visited_count += 1

        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Database error, changes rolled back")

    return UploadResult(
        total=len(parsed_regions),
        visited=visited_count,
        message=f"Updated {len(parsed_regions)} regions, {visited_count} visited",
    )
