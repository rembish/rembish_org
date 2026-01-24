from datetime import date
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..auth.session import get_admin_user
from ..database import get_db
from ..models import Microstate, NMRegion, TCCDestination, UNCountry, User, Visit

router = APIRouter(prefix="/v1/travels", tags=["travels"])


class TravelStats(BaseModel):
    un_visited: int
    un_total: int
    tcc_visited: int
    tcc_total: int
    nm_visited: int
    nm_total: int


class MicrostateData(BaseModel):
    name: str
    longitude: float
    latitude: float
    map_region_code: str


class UNCountryData(BaseModel):
    name: str
    continent: str
    visit_date: str | None  # ISO date or null


class TCCDestinationData(BaseModel):
    name: str
    region: str
    visit_date: str | None  # ISO date or null


class NMRegionData(BaseModel):
    name: str
    country: str
    first_visited_year: int | None
    last_visited_year: int | None


class TravelData(BaseModel):
    stats: TravelStats
    visited_map_regions: dict[str, str]  # region_code -> first_visit_date (ISO)
    visited_countries: list[str]
    microstates: list[MicrostateData]
    un_countries: list[UNCountryData]
    tcc_destinations: list[TCCDestinationData]
    nm_regions: list[NMRegionData]


@router.get("/data", response_model=TravelData)
def get_travel_data(db: Session = Depends(get_db)) -> TravelData:
    """Get travel statistics and visited map regions for the map."""

    # Count UN countries
    un_total = db.query(func.count(UNCountry.id)).scalar() or 0

    # Count visited UN countries (those with at least one visited TCC destination)
    un_visited = (
        db.query(func.count(func.distinct(TCCDestination.un_country_id)))
        .join(Visit, Visit.tcc_destination_id == TCCDestination.id)
        .filter(TCCDestination.un_country_id.isnot(None))
        .scalar()
        or 0
    )

    # Count TCC destinations
    tcc_total = db.query(func.count(TCCDestination.id)).scalar() or 0
    tcc_visited = db.query(func.count(Visit.id)).scalar() or 0

    # Get all map region codes with their first visit dates
    visited_map_regions: dict[str, date] = {}
    visited_countries: list[str] = []

    # Get visited UN countries with earliest visit date
    visited_un_data = (
        db.query(UNCountry, func.min(Visit.first_visit_date).label("first_visit"))
        .join(TCCDestination, TCCDestination.un_country_id == UNCountry.id)
        .join(Visit, Visit.tcc_destination_id == TCCDestination.id)
        .group_by(UNCountry.id)
        .all()
    )

    for country, first_visit in visited_un_data:
        visited_countries.append(country.name)
        # map_region_codes is comma-separated
        for code in country.map_region_codes.split(","):
            code = code.strip()
            if code and first_visit:
                # Keep earliest date if region already exists
                if code not in visited_map_regions or first_visit < visited_map_regions[code]:
                    visited_map_regions[code] = first_visit

    # Get map regions from TCC destinations with their own polygon (e.g., Kosovo, Somaliland)
    visited_tcc_with_polygon = (
        db.query(TCCDestination, Visit.first_visit_date)
        .join(Visit, Visit.tcc_destination_id == TCCDestination.id)
        .filter(
            TCCDestination.map_region_code.isnot(None),
            TCCDestination.un_country_id.is_(None),  # Only non-UN territories
        )
        .all()
    )

    for dest, first_visit in visited_tcc_with_polygon:
        if dest.map_region_code and first_visit:
            code = dest.map_region_code
            if code not in visited_map_regions or first_visit < visited_map_regions[code]:
                visited_map_regions[code] = first_visit
            visited_countries.append(dest.name)

    # Get all microstates
    microstates = db.query(Microstate).all()
    microstates_data = [
        MicrostateData(
            name=m.name,
            longitude=m.longitude,
            latitude=m.latitude,
            map_region_code=m.map_region_code,
        )
        for m in microstates
    ]

    # Convert dates to ISO strings
    visited_map_regions_iso = {code: d.isoformat() for code, d in visited_map_regions.items()}

    # Get all UN countries with their earliest visit date
    all_un_countries = db.query(UNCountry).order_by(UNCountry.continent, UNCountry.name).all()
    un_visit_dates: dict[int, date] = {}
    for country, first_visit in visited_un_data:
        un_visit_dates[country.id] = first_visit

    un_countries_data = [
        UNCountryData(
            name=c.name,
            continent=c.continent,
            visit_date=un_visit_dates[c.id].isoformat() if c.id in un_visit_dates else None,
        )
        for c in all_un_countries
    ]

    # Get all TCC destinations with their visit dates
    all_tcc = (
        db.query(TCCDestination, Visit.first_visit_date)
        .outerjoin(Visit, Visit.tcc_destination_id == TCCDestination.id)
        .order_by(TCCDestination.tcc_region, TCCDestination.name)
        .all()
    )

    tcc_destinations_data = [
        TCCDestinationData(
            name=dest.name,
            region=dest.tcc_region,
            visit_date=visit_date.isoformat() if visit_date else None,
        )
        for dest, visit_date in all_tcc
    ]

    # Get NM stats and regions
    nm_total = db.query(func.count(NMRegion.id)).scalar() or 0
    nm_visited = db.query(func.count(NMRegion.id)).filter(NMRegion.visited.is_(True)).scalar() or 0

    all_nm = db.query(NMRegion).order_by(NMRegion.country, NMRegion.name).all()

    nm_regions_data = [
        NMRegionData(
            name=r.name,
            country=r.country,
            first_visited_year=r.first_visited_year,
            last_visited_year=r.last_visited_year,
        )
        for r in all_nm
    ]

    return TravelData(
        stats=TravelStats(
            un_visited=un_visited,
            un_total=un_total,
            tcc_visited=tcc_visited,
            tcc_total=tcc_total,
            nm_visited=nm_visited,
            nm_total=nm_total,
        ),
        visited_map_regions=visited_map_regions_iso,
        visited_countries=sorted(visited_countries),
        microstates=microstates_data,
        un_countries=un_countries_data,
        tcc_destinations=tcc_destinations_data,
        nm_regions=nm_regions_data,
    )


class UploadResult(BaseModel):
    total: int
    visited: int
    message: str


def extract_country(region_name: str) -> str:
    """Extract country/territory prefix from region name."""
    if region_name.startswith("Timor-Leste"):
        return "Timor-Leste"
    if " – " in region_name:
        return region_name.split(" – ")[0].strip()
    if " - " in region_name:
        return region_name.split(" - ")[0].strip()
    return region_name


@router.post("/upload-nm", response_model=UploadResult)
async def upload_nm_regions(
    file: Annotated[UploadFile, File()],
    admin: Annotated[User, Depends(get_admin_user)],
    db: Session = Depends(get_db),
) -> UploadResult:
    """Upload NomadMania regions XLSX file (admin only)."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be an .xlsx file")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid XLSX file")

    ws = wb.active

    # Parse and validate all data BEFORE touching the database
    parsed_regions: list[dict] = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        name = row[0]
        if not name:
            continue

        try:
            parsed_regions.append(
                {
                    "name": str(name),
                    "country": extract_country(str(name)),
                    "visited": row[1] == 1,
                    "first_visited_year": int(row[2]) if row[2] else None,
                    "last_visited_year": int(row[3]) if row[3] else None,
                }
            )
        except (ValueError, TypeError) as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid data in row {row_num}: {e}",
            )

    if not parsed_regions:
        raise HTTPException(status_code=400, detail="No valid regions found in file")

    # All data validated - now safe to update database
    try:
        db.query(NMRegion).delete()

        visited_count = 0
        for region_data in parsed_regions:
            region = NMRegion(**region_data)
            db.add(region)
            if region_data["visited"]:
                visited_count += 1

        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Database error, changes rolled back")

    return UploadResult(
        total=len(parsed_regions),
        visited=visited_count,
        message=f"Updated {len(parsed_regions)} regions, {visited_count} visited",
    )
